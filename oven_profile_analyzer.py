#!/usr/bin/env python3
"""
炉温曲线分析工具 v1.2
- 导入 GL840 CSV / Snap Cure TXT (GBK) 原始文件
- 解析多通道温度数据，自动检测采样间隔
- 时间-温度曲线图（勾选通道实时刷新，不同颜色）
- 数据间隔提取（线性插值，按秒级间隔提取数据，不可小于原始间隔）
- 导出 Excel（支持勾选转置为横向排列）
- ttkbootstrap flatly 主题，单文件绿色免安装
"""

# ===================== 启动前依赖检查 =====================
import sys
import subprocess
import importlib.util

REQUIRED_PACKAGES = {
    'ttkbootstrap': 'ttkbootstrap',
    'numpy': 'numpy',
    'matplotlib': 'matplotlib',
    'openpyxl': 'openpyxl',
}

def _check_dependencies():
    """检查依赖，缺失时弹出提示并返回缺失列表"""
    missing = []
    for pkg_name, import_name in REQUIRED_PACKAGES.items():
        if importlib.util.find_spec(import_name) is None:
            missing.append(pkg_name)

    if not missing:
        return True

    # 用 tkinter 弹窗提示（Python 自带，必定可用）
    import tkinter as tk
    from tkinter import messagebox
    root = tk.Tk()
    root.withdraw()
    msg = (
        f"缺少以下 Python 依赖包：\n\n"
        f"{' '.join(missing)}\n\n"
        f"请运行以下命令安装：\n"
        f"pip install {' '.join(missing)}\n\n"
        f"如果 pip 不在 PATH 中，请尝试：\n"
        f"python -m pip install {' '.join(missing)}"
    )
    messagebox.showerror("依赖缺失", msg)
    root.destroy()
    return False

if not _check_dependencies():
    sys.exit(1)

# ===================== 正式导入 =====================
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import ttkbootstrap as ttkb
from ttkbootstrap.constants import *
import os
import re
import math
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional

import numpy as np
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
import matplotlib

matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ===================== 常量 =====================
CHANNEL_COLORS = [
    '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd',
    '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf',
    '#aec7e8', '#ffbb78', '#98df8a', '#ff9896', '#c5b0d5',
    '#c49c94', '#f7b6d2', '#c7c7c7', '#dbdb8d', '#9edae5'
]

WINDOW_WIDTH = 1400
WINDOW_HEIGHT = 900
MIN_WIDTH = 1200
MIN_HEIGHT = 800
LEFT_PANEL_WIDTH = 320


# ===================== 数据模型 =====================
@dataclass
class ProfileData:
    """统一的数据模型"""
    file_path: str = ""
    file_type: str = ""  # "gl840" | "snapcure"
    channels: Dict[str, List[Tuple[float, float]]] = field(default_factory=dict)
    # channels: {name: [(time_sec, temperature), ...]}
    original_rate: float = 0.0  # 原始采样率（个/秒）
    metadata: Dict[str, str] = field(default_factory=dict)
    total_points: int = 0
    start_time: Optional[datetime] = None

    @property
    def channel_names(self) -> List[str]:
        return list(self.channels.keys())

    @property
    def time_range(self) -> Tuple[float, float]:
        """返回 (min_time, max_time)"""
        if not self.channels:
            return (0, 0)
        first_ch = next(iter(self.channels.values()))
        if not first_ch:
            return (0, 0)
        return (first_ch[0][0], first_ch[-1][0])

    @property
    def interval_seconds(self) -> float:
        """原始采样间隔（秒）"""
        return 1.0 / self.original_rate if self.original_rate > 0 else 0.0

    @property
    def rate_display(self) -> str:
        """采样间隔显示文本"""
        secs = self.interval_seconds
        if secs <= 0:
            return "-- 秒/次"
        if secs < 60:
            return f"{secs:.1f} 秒/次"
        elif secs < 3600:
            return f"{secs/60:.1f} 分钟/次 ({secs:.0f}秒)"
        else:
            return f"{secs/3600:.1f} 小时/次 ({secs:.0f}秒)"


# ===================== 文件解析器 =====================
def detect_file_type(filepath: str) -> str:
    """自动检测文件类型"""
    # 先尝试 UTF-8
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            head = f.read(500)
            if 'Vendor' in head and ('GL840' in head or 'GL' in head):
                return 'gl840'
            if '曲线开始时间' in head or 'Snap' in head:
                return 'snapcure'
    except UnicodeDecodeError:
        pass

    # 尝试 GBK（Snap Cure 文件）
    try:
        with open(filepath, 'r', encoding='gbk') as f:
            head = f.read(500)
            if '曲线开始时间' in head or 'Snap' in head or '产品名称' in head:
                return 'snapcure'
    except UnicodeDecodeError:
        pass

    # 回退：按扩展名
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        return 'gl840'
    return 'snapcure'


def _parse_sampling_rate(interval_str: str) -> float:
    """解析采样间隔字符串，返回采样率（个/秒）"""
    interval_str = interval_str.strip().lower().replace(' ', '')
    if not interval_str:
        return 0.0

    # 匹配: "1min", "1s", "500ms", "2h" 等
    match = re.match(r'([\d.]+)\s*(min|ms|s|h|sec|m)?', interval_str)
    if not match:
        try:
            return 1.0 / float(interval_str)
        except ValueError:
            return 0.0

    value = float(match.group(1))
    unit = (match.group(2) or 's').lower()

    if unit in ('min', 'm'):
        secs = value * 60
    elif unit == 'ms':
        secs = value / 1000
    elif unit == 'h':
        secs = value * 3600
    else:  # s, sec
        secs = value

    return 1.0 / secs if secs > 0 else 0.0


def parse_gl840(filepath: str) -> ProfileData:
    """解析 GRAPHTEC GL840 CSV 文件"""
    data = ProfileData(file_path=filepath, file_type='gl840')
    channels: Dict[str, List[Tuple[float, float]]] = {}
    metadata = {}

    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    lines = content.split('\n')

    # 解析头部元数据
    in_amp = False
    data_start = -1
    channel_count = 0

    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue

        if line.startswith('Sampling interval'):
            parts = line.split(',')
            if len(parts) >= 2:
                metadata['sampling_interval'] = parts[1].strip()
                data.original_rate = _parse_sampling_rate(parts[1].strip())

        elif line.startswith('Start time'):
            parts = line.split(',')
            if len(parts) >= 3:
                try:
                    time_str = f"{parts[1].strip()} {parts[2].strip()}"
                    data.start_time = datetime.strptime(time_str, '%Y-%m-%d %H:%M:%S')
                except ValueError:
                    pass

        elif line.startswith('Total data points'):
            parts = line.split(',')
            if len(parts) >= 2:
                try:
                    data.total_points = int(parts[1].strip())
                except ValueError:
                    pass

        elif line.startswith('AMP settings'):
            in_amp = True
            continue

        elif in_amp and line.startswith('CH'):
            # CH1," CH 1","M",TEMP,TC_K,... — 跳过表头行 "CH,Signal name,..."
            parts = [p.strip().strip('"') for p in line.split(',')]
            ch_name = parts[0].strip()
            if ch_name == 'CH':
                continue  # 跳过 AMP 表头
            if len(parts) >= 2:
                channels[ch_name] = []
                channel_count += 1

        elif line.startswith('Calc settings'):
            in_amp = False

        elif line.startswith('Data'):
            data_start = i
            break

    if data_start < 0:
        raise ValueError("未找到数据区域（Data 标记）")

    # 解析数据行
    if data.original_rate == 0.0 and data.total_points > 0:
        # 如果头部没有采样间隔，从数据中推断
        data.original_rate = 0.017  # 默认 1min

    # 找到数据标题行
    data_header_idx = -1
    for i in range(data_start + 1, min(data_start + 20, len(lines))):
        if lines[i].strip() and ('NO.' in lines[i] or 'Number' in lines[i]):
            data_header_idx = i
            break

    if data_header_idx < 0:
        data_header_idx = data_start + 1

    # 解析数据行
    ch_names = sorted(channels.keys(), key=lambda x: int(re.search(r'\d+', x).group()))
    first_time = None

    for i in range(data_header_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line or line.startswith('*') or line.startswith('#'):
            continue

        parts = [p.strip() for p in line.split(',')]
        if len(parts) < 4 + channel_count:
            continue

        try:
            # 解析时间（Date&Time 是单列: "2026/06/26 11:36:54"）
            datetime_str = parts[1].strip()
            dt = datetime.strptime(datetime_str, '%Y/%m/%d %H:%M:%S')
            if first_time is None:
                first_time = dt
            elapsed = (dt - first_time).total_seconds()

            # 解析通道数据（parts[3] 开始是 CH1, CH2, ...）
            for j, ch_name in enumerate(ch_names):
                idx = 3 + j
                if idx < len(parts):
                    val_str = parts[idx].replace('+', '').replace(' ', '')
                    if val_str and val_str not in ('', '-', '--'):
                        temp = float(val_str)
                        channels[ch_name].append((elapsed, temp))

        except (ValueError, IndexError):
            continue

    # 过滤空通道
    data.channels = {k: v for k, v in channels.items() if v}
    data.metadata = metadata

    # 如果通道数量与数据不匹配，从实际数据行重新推断
    if not data.channels:
        actual_ch_count = 0
        for i in range(data_header_idx + 1, min(data_header_idx + 5, len(lines))):
            parts = [p.strip() for p in lines[i].split(',')]
            if len(parts) > 4:
                # 计算实际温度列数
                temp_count = 0
                for p in parts[3:]:
                    if p and ('+' in p or '-' in p) and not p.startswith('L') and not p.startswith('A'):
                        temp_count += 1
                    else:
                        break
                actual_ch_count = max(actual_ch_count, temp_count)

        if actual_ch_count > 0:
            data.channels = {}
            for j in range(actual_ch_count):
                ch_name = f"CH{j + 1}"
                data.channels[ch_name] = []

            first_time = None
            for i in range(data_header_idx + 1, len(lines)):
                line = lines[i].strip()
                if not line:
                    continue
                parts = [p.strip() for p in line.split(',')]
                if len(parts) < 4 + actual_ch_count:
                    continue
                try:
                    datetime_str = parts[1].strip()
                    dt = datetime.strptime(datetime_str, '%Y/%m/%d %H:%M:%S')
                    if first_time is None:
                        first_time = dt
                    elapsed = (dt - first_time).total_seconds()
                    for j in range(actual_ch_count):
                        idx = 3 + j
                        if idx < len(parts):
                            val_str = parts[idx].replace('+', '').replace(' ', '')
                            if val_str and val_str not in ('', '-', '--'):
                                data.channels[f"CH{j + 1}"].append((elapsed, float(val_str)))
                except (ValueError, IndexError):
                    continue

    data.total_points = max((len(v) for v in data.channels.values()), default=0)
    return data


def parse_snapcure(filepath: str) -> ProfileData:
    """解析 Snap Cure 回流焊 TXT 文件（GBK 编码）"""
    data = ProfileData(file_path=filepath, file_type='snapcure')
    metadata = {}

    with open(filepath, 'r', encoding='gbk', errors='replace') as f:
        content = f.read()

    lines = content.split('\n')

    # 解析头部元数据
    for line in lines[:80]:
        line = line.strip()
        if '产品名称' in line:
            parts = line.split('\t')
            metadata['product'] = parts[-1].strip() if len(parts) > 1 else ''
        elif '制程界限名称' in line:
            parts = line.split('\t')
            metadata['process_limit'] = parts[-1].strip() if len(parts) > 1 else ''
        elif '有效热电偶' in line:
            parts = line.split('\t')
            metadata['active_tc'] = parts[-1].strip() if len(parts) > 1 else ''

    # 查找数据区域："未经处理" 段落
    data_start = -1
    for i, line in enumerate(lines):
        if '未经处理' in line:
            data_start = i
            break

    if data_start < 0:
        # 尝试直接找 "秒\tTC1" 行
        for i, line in enumerate(lines):
            if '秒' in line and 'TC1' in line:
                data_start = i - 1
                break

    if data_start < 0:
        raise ValueError("未找到数据区域（'未经处理' 标记）")

    # 找数据标题行
    header_idx = -1
    tc_count = 0
    for i in range(data_start, min(data_start + 30, len(lines))):
        line = lines[i].strip()
        if '秒' in line and 'TC1' in line:
            header_idx = i
            # 统计原始数据区 TC 列数（只数第一个 "秒" 后的 TC 列，重组数据区不算）
            parts = line.split('\t')
            tc_count = 0
            for p in parts:
                p = p.strip()
                if p == '秒':
                    if tc_count > 0:
                        break  # 遇到第二个"秒"，停止计数
                    continue
                if p.startswith('TC'):
                    tc_count += 1
            break

    if header_idx < 0:
        raise ValueError("未找到数据标题行")

    # 初始化通道
    channels: Dict[str, List[Tuple[float, float]]] = {}
    for j in range(tc_count):
        channels[f"TC{j + 1}"] = []

    # 解析数据行
    for i in range(header_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line:
            continue

        parts = line.split('\t')
        # 清理空元素
        parts = [p.strip() for p in parts if p.strip()]

        if len(parts) < 1 + tc_count:
            continue

        try:
            time_sec = float(parts[0])
            for j in range(tc_count):
                if j + 1 < len(parts):
                    val = float(parts[j + 1])
                    channels[f"TC{j + 1}"].append((time_sec, val))
        except (ValueError, IndexError):
            continue

    data.channels = {k: v for k, v in channels.items() if v}
    data.total_points = max((len(v) for v in data.channels.values()), default=0)

    # 计算采样率
    if data.total_points > 1:
        first_ch = next(iter(data.channels.values()))
        if len(first_ch) >= 2:
            time_span = first_ch[-1][0] - first_ch[0][0]
            if time_span > 0:
                data.original_rate = (data.total_points - 1) / time_span

    data.metadata = metadata
    return data


def parse_file(filepath: str) -> ProfileData:
    """统一入口：自动检测文件类型并解析"""
    file_type = detect_file_type(filepath)
    if file_type == 'gl840':
        return parse_gl840(filepath)
    else:
        return parse_snapcure(filepath)


# ===================== 重采样引擎 =====================
def resample_data(data: ProfileData, target_rate: float) -> ProfileData:
    """
    线性插值重采样
    返回新的 ProfileData，包含重采样后的数据
    """
    if target_rate <= 0 or target_rate > data.original_rate:
        raise ValueError(f"目标采样率必须在 0 ~ {data.original_rate:.4f} 个/秒之间")

    interval = 1.0 / target_rate
    t_min, t_max = data.time_range
    new_times = np.arange(t_min, t_max + interval * 0.5, interval)

    new_channels = {}
    for ch_name, points in data.channels.items():
        if not points:
            continue
        old_times = np.array([p[0] for p in points])
        old_temps = np.array([p[1] for p in points])
        new_temps = np.interp(new_times, old_times, old_temps)
        new_channels[ch_name] = list(zip(new_times.tolist(), new_temps.tolist()))

    result = ProfileData(
        file_path=data.file_path,
        file_type=data.file_type,
        channels=new_channels,
        original_rate=target_rate,
        metadata=data.metadata.copy(),
        total_points=len(new_times),
        start_time=data.start_time
    )
    return result


# ===================== Excel 导出 =====================
def export_excel(data: ProfileData, output_path: str, channel_names: List[str],
                 transpose: bool = False):
    """导出为 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "炉温数据"

    # 样式
    header_font = Font(name='Microsoft YaHei', bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
    data_font = Font(name='Microsoft YaHei', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')

    # 获取数据
    selected_channels = {k: data.channels[k] for k in channel_names if k in data.channels}
    if not selected_channels:
        return

    if not transpose:
        # 纵向排列：行=时间点，列=通道
        # 获取所有时间点（取第一个通道的时间轴）
        first_ch = next(iter(selected_channels.values()))
        time_points = [p[0] for p in first_ch]
        num_points = len(time_points)

        # 写标题行
        ws.cell(row=1, column=1, value='时间 (秒)').font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).border = thin_border

        for j, ch_name in enumerate(channel_names):
            if ch_name in selected_channels:
                cell = ws.cell(row=1, column=j + 2, value=f'{ch_name} (°C)')
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

        # 写数据
        for i in range(num_points):
            time_val = time_points[i]
            ws.cell(row=i + 2, column=1, value=round(time_val, 2)).font = data_font
            ws.cell(row=i + 2, column=1).alignment = center_align
            ws.cell(row=i + 2, column=1).border = thin_border

            col = 2
            for ch_name in channel_names:
                if ch_name in selected_channels:
                    points = selected_channels[ch_name]
                    if i < len(points):
                        temp = round(points[i][1], 2)
                        ws.cell(row=i + 2, column=col, value=temp).font = data_font
                        ws.cell(row=i + 2, column=col).alignment = center_align
                        ws.cell(row=i + 2, column=col).border = thin_border
                    col += 1

        # 冻结首行
        ws.freeze_panes = 'B2'

    else:
        # 横向排列：行=通道，列=时间点
        first_ch = next(iter(selected_channels.values()))
        time_points = [p[0] for p in first_ch]
        num_points = len(time_points)

        # 写标题行（时间点）
        ws.cell(row=1, column=1, value='通道').font = header_font_white
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = center_align
        ws.cell(row=1, column=1).border = thin_border

        for j, t in enumerate(time_points):
            cell = ws.cell(row=1, column=j + 2, value=round(t, 2))
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 写数据
        row = 2
        for ch_name in channel_names:
            if ch_name not in selected_channels:
                continue
            points = selected_channels[ch_name]

            ws.cell(row=row, column=1, value=f'{ch_name} (°C)').font = header_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=1).border = thin_border

            for j in range(num_points):
                if j < len(points):
                    temp = round(points[j][1], 2)
                    ws.cell(row=row, column=j + 2, value=temp).font = data_font
                    ws.cell(row=row, column=j + 2).alignment = center_align
                    ws.cell(row=row, column=j + 2).border = thin_border
            row += 1

        # 冻结首行首列
        ws.freeze_panes = 'B2'

    # 自动列宽
    for col_idx in range(1, ws.max_column + 1):
        max_width = 0
        for row_idx in range(1, min(ws.max_row + 1, 100)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                width = len(str(cell.value)) * 1.3 + 2
                max_width = max(max_width, min(width, 25))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(8, max_width)

    wb.save(output_path)


# ===================== 主应用 =====================
class OvenProfileApp:
    def __init__(self, root):
        self.root = root
        self.root.title("炉温曲线分析工具 v1.2")
        self.root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
        self.root.minsize(MIN_WIDTH, MIN_HEIGHT)

        # 数据状态
        self.original_data: Optional[ProfileData] = None
        self.resampled_data: Optional[ProfileData] = None
        self.current_data: Optional[ProfileData] = None  # 当前显示的数据
        self.channel_vars: Dict[str, tk.BooleanVar] = {}
        self.transpose_var = tk.BooleanVar(value=False)

        # 图表引用
        self.figure: Optional[Figure] = None
        self.canvas: Optional[FigureCanvasTkAgg] = None
        self.ax = None

        self._setup_ui()

    # ===================== UI 搭建 =====================
    def _setup_ui(self):
        # 顶部栏
        self._build_top_bar()

        # 垂直分割：上方（左侧面板+图表） | 下方（数据预览）
        self.v_paned = ttkb.Panedwindow(self.root, orient=VERTICAL)
        self.v_paned.pack(fill=BOTH, expand=True, padx=4, pady=(0, 4))

        # 上半部分：水平分割（左侧面板 + 图表）
        self.paned = ttkb.Panedwindow(self.v_paned, orient=HORIZONTAL)
        self.v_paned.add(self.paned, weight=7)

        # 左面板
        self.left_frame = ttkb.Frame(self.paned, width=LEFT_PANEL_WIDTH)
        self.paned.add(self.left_frame, weight=0)
        self._build_left_panel()

        # 右面板（图表）
        self.right_frame = ttkb.Frame(self.paned)
        self.paned.add(self.right_frame, weight=1)
        self._build_chart_area()

        # 底部面板（数据预览）
        self._build_bottom_panel()

    def _build_top_bar(self):
        top = ttkb.Frame(self.root, padding=6)
        top.pack(fill=X, padx=4, pady=(4, 0))

        self.btn_import = ttkb.Button(top, text="📂 导入文件", bootstyle="primary",
                                       command=self._import_file, width=14)
        self.btn_import.pack(side=LEFT, padx=(0, 12))

        self.lbl_file = ttkb.Label(top, text="就绪，请导入原始文件", font=("Microsoft YaHei", 10))
        self.lbl_file.pack(side=LEFT, padx=4)

        self.lbl_info = ttkb.Label(top, text="", font=("Microsoft YaHei", 9),
                                    foreground="gray")
        self.lbl_info.pack(side=RIGHT, padx=4)

    def _build_left_panel(self):
        # 通道选择
        self.ch_frame = ttkb.Labelframe(self.left_frame, text="通道选择", padding=6)
        self.ch_frame.pack(fill=X, padx=2, pady=(0, 4))

        btn_row = ttkb.Frame(self.ch_frame)
        btn_row.pack(fill=X, pady=(0, 4))
        ttkb.Button(btn_row, text="全选", bootstyle="secondary-outline",
                     command=self._on_select_all, width=8).pack(side=LEFT, padx=2)
        ttkb.Button(btn_row, text="反选", bootstyle="secondary-outline",
                     command=self._on_invert, width=8).pack(side=LEFT, padx=2)

        # 通道复选框容器（可滚动）— height=180 确保 18 通道 6 行全部可见
        self.ch_canvas = tk.Canvas(self.ch_frame, height=180, highlightthickness=0)
        self.ch_scrollbar = ttkb.Scrollbar(self.ch_frame, orient=VERTICAL, command=self.ch_canvas.yview)
        self.ch_inner = ttkb.Frame(self.ch_canvas)
        self.ch_inner.bind('<Configure>', lambda e: self.ch_canvas.configure(scrollregion=self.ch_canvas.bbox('all')))
        self.ch_canvas.create_window((0, 0), window=self.ch_inner, anchor='nw')
        self.ch_canvas.configure(yscrollcommand=self.ch_scrollbar.set)
        self.ch_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        self.ch_scrollbar.pack(side=RIGHT, fill=Y)
        # 鼠标滚轮支持 — 绑定到 canvas 和 inner 上，确保在通道上也能滚动
        self.ch_canvas.bind('<MouseWheel>', self._on_ch_mousewheel)
        self.ch_inner.bind('<MouseWheel>', self._on_ch_mousewheel)

        # 数据间隔提取
        self.resample_frame = ttkb.Labelframe(self.left_frame, text="数据间隔提取", padding=6)
        self.resample_frame.pack(fill=X, padx=2, pady=4)

        # 间隔单位变量
        self.interval_unit_var = tk.StringVar(value="s")

        self.lbl_orig_rate = ttkb.Label(self.resample_frame,
                                          text="原始间隔: -- 秒/次",
                                          font=("Microsoft YaHei", 9))
        self.lbl_orig_rate.pack(anchor=W, pady=(0, 4))

        row = ttkb.Frame(self.resample_frame)
        row.pack(fill=X)
        ttkb.Label(row, text="每隔", font=("Microsoft YaHei", 9)).pack(side=LEFT)
        self.entry_target_rate = ttkb.Entry(row, width=8, font=("Microsoft YaHei", 9))
        self.entry_target_rate.pack(side=LEFT, padx=2)
        self.combo_unit = ttkb.Combobox(row, textvariable=self.interval_unit_var,
                                         values=["s (秒)", "min (分)"],
                                         width=10, font=("Microsoft YaHei", 9),
                                         state="readonly")
        self.combo_unit.pack(side=LEFT, padx=2)
        ttkb.Label(row, text="取一个点", font=("Microsoft YaHei", 9)).pack(side=LEFT)
        self.combo_unit.bind('<<ComboboxSelected>>', lambda e: self._on_interval_input())
        self.entry_target_rate.bind('<Key>', lambda e: self.entry_target_rate.configure(foreground=''))
        self.entry_target_rate.bind('<KeyRelease>', self._on_interval_input)

        self.lbl_rate_hint = ttkb.Label(self.resample_frame, text="",
                                          font=("Microsoft YaHei", 8), foreground="gray")
        self.lbl_rate_hint.pack(anchor=W, pady=(2, 0))

        # 快捷间隔按钮
        quick_row = ttkb.Frame(self.resample_frame)
        quick_row.pack(fill=X, pady=(4, 0))
        quick_intervals = [1, 5, 10, 30, 60, 120]
        for qi in quick_intervals:
            btn = ttkb.Button(quick_row, text=f"{qi}s", bootstyle="secondary-outline",
                              width=4, command=lambda v=qi: self._set_interval(v))
            btn.pack(side=LEFT, padx=1)

        btn_row2 = ttkb.Frame(self.resample_frame)
        btn_row2.pack(fill=X, pady=(6, 0))
        self.btn_apply = ttkb.Button(btn_row2, text="提取数据", bootstyle="success",
                                      command=self._apply_resample, state=DISABLED)
        self.btn_apply.pack(side=LEFT, padx=(0, 4))
        self.btn_reset = ttkb.Button(btn_row2, text="重置", bootstyle="secondary-outline",
                                      command=self._reset_resample, state=DISABLED)
        self.btn_reset.pack(side=LEFT)

        self.lbl_resample_status = ttkb.Label(self.resample_frame, text="",
                                                font=("Microsoft YaHei", 8), foreground="green")
        self.lbl_resample_status.pack(anchor=W, pady=(4, 0))

    def _build_chart_area(self):
        """图表区域"""
        self.fig = Figure(figsize=(8, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.ax.set_xlabel("时间 (秒)", fontsize=10)
        self.ax.set_ylabel("温度 (°C)", fontsize=10)
        self.ax.grid(True, alpha=0.3)
        self.ax.set_title("请导入文件", fontsize=12)
        self.fig.tight_layout()

        # Canvas
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.right_frame)
        self.canvas.get_tk_widget().pack(fill=BOTH, expand=True)

        # 工具栏
        self.toolbar = NavigationToolbar2Tk(self.canvas, self.right_frame)
        self.toolbar.update()

    def _build_bottom_panel(self):
        """底部：数据表格 + 导出"""
        bottom = ttkb.Labelframe(self.v_paned, text="数据预览", padding=4)
        self.v_paned.add(bottom, weight=1)

        # Treeview 容器
        table_frame = ttkb.Frame(bottom)
        table_frame.pack(fill=BOTH, expand=True)

        # 水平滚动条
        self.table_scroll_x = ttkb.Scrollbar(table_frame, orient=HORIZONTAL)
        self.table_scroll_y = ttkb.Scrollbar(table_frame, orient=VERTICAL)

        self.tree = ttkb.Treeview(table_frame, height=6, show='headings',
                                   xscrollcommand=self.table_scroll_x.set,
                                   yscrollcommand=self.table_scroll_y.set)
        self.table_scroll_x.config(command=self.tree.xview)
        self.table_scroll_y.config(command=self.tree.yview)

        self.tree.grid(row=0, column=0, sticky='nsew')
        self.table_scroll_y.grid(row=0, column=1, sticky='ns')
        self.table_scroll_x.grid(row=1, column=0, sticky='ew')
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # 导出栏 — side=BOTTOM 确保拉伸窗口时按钮始终可见
        export_row = ttkb.Frame(bottom)
        export_row.pack(fill=X, side=BOTTOM, pady=(6, 0))

        self.chk_transpose = ttkb.Checkbutton(export_row, text="转置横向排列（行=通道，列=时间点）",
                                               variable=self.transpose_var,
                                               bootstyle="info-round-toggle")
        self.chk_transpose.pack(side=LEFT, padx=4)

        self.btn_export = ttkb.Button(export_row, text="📊 导出 Excel", bootstyle="info",
                                       command=self._export_excel, state=DISABLED, width=16)
        self.btn_export.pack(side=RIGHT, padx=4)

        self.lbl_table_status = ttkb.Label(export_row, text="",
                                             font=("Microsoft YaHei", 9), foreground="gray")
        self.lbl_table_status.pack(side=RIGHT, padx=10)

    # ===================== 交互逻辑 =====================
    def _import_file(self):
        """导入文件"""
        filepath = filedialog.askopenfilename(
            title="选择原始文件",
            filetypes=[("原始文件", "*.csv *.txt"), ("CSV 文件", "*.csv"), ("TXT 文件", "*.txt"),
                       ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            self.original_data = parse_file(filepath)
            self.current_data = self.original_data
            self.resampled_data = None

            # 更新 UI
            fname = os.path.basename(filepath)
            ch_count = len(self.original_data.channel_names)
            self.lbl_file.config(text=f"文件: {fname}")
            self.lbl_info.config(
                text=f"通道: {ch_count} | 采样: {self.original_data.rate_display} | "
                     f"数据点: {self.original_data.total_points}"
            )

            # 重建通道复选框
            self._rebuild_channel_checkboxes()

            # 更新采样率显示
            self.lbl_orig_rate.config(text=f"原始间隔: {self.original_data.rate_display}")
            self.entry_target_rate.delete(0, tk.END)
            orig_secs = self.original_data.interval_seconds
            orig_mins = orig_secs / 60
            self.lbl_rate_hint.config(text=f"不小于原始间隔: {orig_secs:.1f} 秒 ({orig_mins:.2f} 分)")
            self.lbl_resample_status.config(text="")
            self.btn_apply.config(state=NORMAL)
            self.btn_reset.config(state=DISABLED)
            self.btn_export.config(state=NORMAL)

            # 更新图表和表格
            self._update_chart()
            self._update_table()

        except Exception as e:
            messagebox.showerror("解析错误", f"文件解析失败:\n{str(e)}")

    def _rebuild_channel_checkboxes(self):
        """重建通道复选框"""
        for w in self.ch_inner.winfo_children():
            w.destroy()
        self.channel_vars.clear()

        if not self.original_data:
            return

        ch_names = self.original_data.channel_names
        # 每行最多 3 个
        for i, name in enumerate(ch_names):
            var = tk.BooleanVar(value=True)
            self.channel_vars[name] = var
            cb = ttkb.Checkbutton(self.ch_inner, text=name, variable=var,
                                   bootstyle="info-round-toggle",
                                   command=self._on_channel_toggle)
            cb.grid(row=i // 3, column=i % 3, sticky=W, padx=2, pady=1)

    def _update_chart(self):
        """更新图表"""
        self.ax.clear()

        if not self.current_data or not self.current_data.channels:
            self.ax.set_title("请导入文件", fontsize=12)
            self.canvas.draw()
            return

        data = self.current_data
        is_resampled = (self.resampled_data is not None and
                        self.current_data is self.resampled_data)

        # 只绘制勾选的通道
        selected = self._get_selected_channels()
        if not selected:
            self.ax.set_title("未选择通道", fontsize=12)
            self.canvas.draw()
            return

        for i, ch_name in enumerate(selected):
            points = data.channels[ch_name]
            if not points:
                continue
            times = [p[0] for p in points]
            temps = [p[1] for p in points]
            color = CHANNEL_COLORS[i % len(CHANNEL_COLORS)]

            if is_resampled:
                # 重采样数据：带标记点
                self.ax.plot(times, temps, color=color, linewidth=1.0,
                            marker='o', markersize=2, markevery=max(1, len(times) // 50),
                            label=ch_name)
            else:
                # 原始数据：细线
                self.ax.plot(times, temps, color=color, linewidth=1.0, label=ch_name)

        self.ax.set_xlabel("时间 (秒)", fontsize=10)
        self.ax.set_ylabel("温度 (°C)", fontsize=10)
        self.ax.grid(True, alpha=0.3)

        title = os.path.basename(self.original_data.file_path) if self.original_data else ""
        if is_resampled:
            title += f"  [间隔: {self.resampled_data.interval_seconds:.1f}秒/次]"
        self.ax.set_title(title, fontsize=11)

        # 图例 — 始终放在图表外部，避免遮蔽数据曲线
        n_selected = len(selected)
        self.ax.legend(fontsize=8, loc='upper left',
                      bbox_to_anchor=(1.01, 1), borderaxespad=0)

        self.fig.subplots_adjust(left=0.08, right=0.80, top=0.92, bottom=0.08)
        self.canvas.draw()

    def _update_table(self):
        """更新数据表格"""
        self.tree.delete(*self.tree.get_children())

        # 清除旧列
        cols = self.tree['columns']
        if cols:
            self.tree['columns'] = []

        if not self.current_data:
            self.lbl_table_status.config(text="")
            return

        # 获取选中的通道
        selected = self._get_selected_channels()
        if not selected:
            self.lbl_table_status.config(text="未选择通道")
            return

        # 设置列
        columns = ['time'] + selected
        self.tree['columns'] = columns
        self.tree.heading('time', text='时间 (秒)')
        self.tree.column('time', width=90, anchor='center', minwidth=70)

        for ch in selected:
            self.tree.heading(ch, text=f'{ch} (°C)')
            self.tree.column(ch, width=85, anchor='center', minwidth=70)

        # 填充数据
        first_ch = self.current_data.channels.get(selected[0], [])
        num_rows = len(first_ch)
        # 限制显示行数
        display_rows = min(num_rows, 500)

        for i in range(display_rows):
            values = []
            for ch in selected:
                points = self.current_data.channels.get(ch, [])
                if i < len(points):
                    temp_val = f"{points[i][1]:.2f}"
                    values.append(temp_val)
                else:
                    values.append('')
            row_values = [f"{first_ch[i][0]:.2f}" if i < len(first_ch) else ""] + values
            self.tree.insert('', 'end', values=row_values)

        total = num_rows
        shown = display_rows
        hint = f"显示 {shown}/{total} 行" if shown < total else f"共 {total} 行"
        self.lbl_table_status.config(text=hint)

    def _get_selected_channels(self) -> List[str]:
        """获取勾选的通道名称列表"""
        if not self.original_data:
            return []
        return [name for name in self.original_data.channel_names
                if self.channel_vars.get(name, tk.BooleanVar(value=True)).get()]

    def _on_interval_input(self, event=None):
        """输入间隔时实时预览预计数据点"""
        try:
            val = self.entry_target_rate.get().strip()
            if not val:
                orig_secs = self.original_data.interval_seconds
                orig_mins = orig_secs / 60
                self.lbl_rate_hint.config(text=f"不小于原始间隔: {orig_secs:.1f} 秒 ({orig_mins:.2f} 分)")
                return
            target_val = float(val)
            if target_val <= 0:
                return
            # 根据单位转换为秒
            unit = self.interval_unit_var.get()
            if unit.startswith("min"):
                target_interval = target_val * 60
                unit_label = "分"
            else:
                target_interval = target_val
                unit_label = "秒"
            orig_interval = self.original_data.interval_seconds
            if target_interval < orig_interval:
                self.lbl_rate_hint.config(
                    text=f"⚠ 不可小于原始间隔 {orig_interval:.1f} 秒 ({orig_interval/60:.1f}分)",
                    foreground="red")
                return
            est_points = int(self.original_data.total_points * orig_interval / target_interval) + 1
            self.lbl_rate_hint.config(
                text=f"预计约 {est_points} 个数据点 ({target_val}{unit_label}×{self.original_data.total_points}点)",
                foreground="gray")
        except ValueError:
            pass

    def _set_interval(self, seconds):
        """快捷设置间隔（始终用秒单位）"""
        self.interval_unit_var.set("s (秒)")
        self.entry_target_rate.delete(0, tk.END)
        self.entry_target_rate.insert(0, str(seconds))
        self._on_interval_input()

    def _on_ch_mousewheel(self, event):
        """通道区域鼠标滚轮"""
        self.ch_canvas.yview_scroll(int(-1 * (event.delta / 120)), 'units')

    def _apply_resample(self):
        """按时间间隔提取数据"""
        if not self.original_data:
            return

        try:
            target_str = self.entry_target_rate.get().strip()
            if not target_str:
                messagebox.showwarning("提示", "请输入时间间隔")
                return
            target_val = float(target_str)
            # 根据单位转换为秒
            unit = self.interval_unit_var.get()
            if unit.startswith("min"):
                target_interval = target_val * 60
                unit_display = f"{target_val}分 ({target_interval:.0f}秒)"
            else:
                target_interval = target_val
                unit_display = f"{target_val}秒"
        except ValueError:
            messagebox.showwarning("提示", "请输入有效的数字")
            self.entry_target_rate.configure(foreground='red')
            return

        orig_interval = self.original_data.interval_seconds

        if target_interval <= 0:
            messagebox.showwarning("提示", "间隔必须大于 0")
            self.entry_target_rate.configure(foreground='red')
            return

        if target_interval < orig_interval:
            messagebox.showwarning("提示",
                                     f"目标间隔不能小于原始间隔\n"
                                     f"原始间隔: {orig_interval:.1f} 秒 ({orig_interval/60:.1f} 分)\n"
                                     f"请输入 ≥ {orig_interval:.1f} 秒")
            self.entry_target_rate.configure(foreground='red')
            return

        self.entry_target_rate.configure(foreground='')

        # 内部转换为采样率
        target_rate = 1.0 / target_interval

        try:
            self.resampled_data = resample_data(self.original_data, target_rate)
            self.current_data = self.resampled_data

            new_interval = self.resampled_data.interval_seconds
            new_points = self.resampled_data.total_points
            self.lbl_resample_status.config(
                text=f"✓ 每 {new_interval:.1f} 秒取一个点（{unit_display}），共 {new_points} 个数据点"
            )
            self.btn_reset.config(state=NORMAL)

            self._update_chart()
            self._update_table()

        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _reset_resample(self):
        """重置重采样"""
        self.resampled_data = None
        self.current_data = self.original_data
        self.entry_target_rate.delete(0, tk.END)
        self.lbl_resample_status.config(text="")
        self.btn_reset.config(state=DISABLED)
        self._update_chart()
        self._update_table()

    def _export_excel(self):
        """导出 Excel"""
        if not self.current_data:
            return

        selected = self._get_selected_channels()
        if not selected:
            messagebox.showwarning("提示", "请至少选择一个通道")
            return

        # 生成默认文件名
        orig_name = os.path.splitext(os.path.basename(
            self.original_data.file_path if self.original_data else "export"
        ))[0]
        rate_suffix = ""
        if self.resampled_data:
            rate_suffix = f"_interval_{self.resampled_data.interval_seconds:.0f}s"
        default_name = f"{orig_name}{rate_suffix}.xlsx"

        filepath = filedialog.asksaveasfilename(
            title="导出 Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not filepath:
            return

        try:
            export_excel(self.current_data, filepath, selected,
                         transpose=self.transpose_var.get())
            messagebox.showinfo("导出成功", f"文件已保存:\n{filepath}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def _on_channel_toggle(self):
        """通道勾选变化时更新图表和表格"""
        self._update_chart()
        self._update_table()

    def _on_select_all(self):
        """全选通道"""
        for var in self.channel_vars.values():
            var.set(True)
        self._update_chart()
        self._update_table()

    def _on_invert(self):
        """反选通道"""
        for var in self.channel_vars.values():
            var.set(not var.get())
        self._update_chart()
        self._update_table()


# ===================== 入口 =====================
def main():
    root = ttkb.Window(themename="flatly")
    app = OvenProfileApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
